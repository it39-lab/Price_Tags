from itertools import product
from tkinter import N
import openpyxl

products_file = openpyxl.load_workbook("Products.xlsx")
ws = products_file["Sheet1"]

product_list = {}

for product_row in range(2, ws.max_row + 1):
    product_name = ws.cell(product_row, 1).value
    product_price = ws.cell(product_row, 2).value
    product_list.update({product_name: product_price})
print(product_list)
